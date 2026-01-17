import re
from openpyxl import load_workbook

#Logica de negocio 
def clean_id(value):
    # Elimina caracteres no numericos de un documento
    if value is None: 
        return ""
    return re.sub(r"\D",'',str(value))  #funcion que lo hace

def  merge_name(name, lastname):
    if name is None:
        name = ""
    if name is None:
       lastname = ""   
    return f"{name} {lastname}".strip()

def process_excel(path):
     #Acceso al ex
     wb = load_workbook(path)
     ws= wb ["Hoja 1"]
     for row in range (2,ws.max_row+1):  
         ws[f"D{row}"]=clean_id(ws[f"A{row}"].value)
         #cedula
         #nombre 
         ws[f"E{row}"]=merge_name(
         ws[f"B{row}"].value,
         ws[f"C{row}"].value
         )    
     wb.save(path)  #guardar cambios realizados en el excel
 #Controlador
def process_excel_safe(path):
    try:
        process_excel(path)
        return True, "Archivo Procesado Correctamente"
    except PermissionError:
        return(
            False,
            "El archivo excel esta abierto.\n"
            "porfavor, Cierrelo e intentelo de uno"
        )
    except KeyError:
        return False,"Hojas 'Datos' no encontrada"
    except Exception as e:
        return False,f"Error inesperado :{str(e)}"
        